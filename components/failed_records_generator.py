import streamlit as st
import pandas as pd
import json
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.ge_helpers import GEHelpers
from config.app_config import AppConfig

class FailedRecordsGenerator:
    """Component for generating downloadable reports of failed records"""
    
    def __init__(self):
        self.ge_helpers = GEHelpers()
        self.config = AppConfig()
    
    def render(self, validation_results: Dict, original_data: pd.DataFrame):
        """Render the failed records generator interface"""
        st.markdown("### 📊 Failed Records Report Generator")
        st.markdown("Generate a comprehensive report of all records that failed validation expectations.")
        
        if not validation_results:
            st.error("No validation results available! Please run validation first.")
            return
        
        if original_data is None:
            st.error("Original data not available! Cannot generate failed records report.")
            return
        
        # Show summary of what will be generated
        self._render_generation_summary(validation_results, original_data)
        
        # Generation options
        self._render_generation_options()
        
        # Generate and download button
        self._render_generate_button(validation_results, original_data)
        
        # Show preview if available
        if 'failed_records_data' in st.session_state:
            self._render_failed_records_preview()
    
    def _render_generation_summary(self, validation_results: Dict, original_data: pd.DataFrame):
        """Render summary of what will be generated"""
        st.markdown("#### 📋 What You'll Get")
        
        # Calculate failed records statistics
        failed_stats = self._calculate_failed_records_stats(validation_results, original_data)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Failed Records", f"{failed_stats['total_failed_records']:,}")
        
        with col2:
            st.metric("Failed Expectations", failed_stats['failed_expectations'])
        
        with col3:
            st.metric("Affected Columns", failed_stats['affected_columns'])
        
        with col4:
            if failed_stats['total_failed_records'] > 0:
                failure_rate = (failed_stats['total_failed_records'] / len(original_data)) * 100
                st.metric("Overall Failure Rate", f"{failure_rate:.1f}%")
            else:
                st.metric("Overall Failure Rate", "0.0%")
        
        # Show breakdown by column
        if failed_stats['breakdown_by_type']:
            st.markdown("**Breakdown by Column:**")
            breakdown_df = pd.DataFrame(failed_stats['breakdown_by_type'])
            st.dataframe(
                breakdown_df,
                use_container_width=True,
                column_config={
                    "Column": st.column_config.TextColumn("Column", width="medium"),
                    "Failed Records": st.column_config.NumberColumn("Failed Records", width="small"),
                    "Failure Rate": st.column_config.TextColumn("Failure Rate", width="small")
                }
            )
    
    def _render_generation_options(self):
        """Render generation options"""
        st.markdown("#### ⚙️ Report Options")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Report format
            report_format = st.selectbox(
                "Report Format:",
                options=["CSV", "Excel", "JSON"],
                help="Choose the format for your failed records report"
            )
            st.session_state.failed_records_format = report_format
            
            # Include metadata
            include_metadata = st.checkbox(
                "Include validation metadata",
                value=True,
                help="Include expectation details, failure reasons, and validation context"
            )
            st.session_state.include_metadata = include_metadata
        
        with col2:
            # Include original data
            include_original = st.checkbox(
                "Include original record data",
                value=True,
                help="Include the complete original record data for failed records"
            )
            st.session_state.include_original = include_original
            
            # Group by expectation
            group_by_expectation = st.checkbox(
                "Group by expectation type",
                value=False,
                help="Organize failed records by the expectation that failed"
            )
            st.session_state.group_by_expectation = group_by_expectation
        
        # Additional options
        st.markdown("**Additional Options:**")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Max records per file (for large datasets)
            max_records = st.number_input(
                "Max records per file:",
                min_value=1000,
                max_value=100000,
                value=50000,
                step=1000,
                help="Split large reports into multiple files for easier handling"
            )
            st.session_state.max_records_per_file = max_records
        
        with col2:
            # Include success summary
            include_success_summary = st.checkbox(
                "Include success summary",
                value=True,
                help="Add a summary of passed expectations for context"
            )
            st.session_state.include_success_summary = include_success_summary
        
        with col3:
            # Timestamp format
            timestamp_format = st.selectbox(
                "Timestamp format:",
                options=["ISO", "Readable", "Unix"],
                help="Choose how to format timestamps in the report"
            )
            st.session_state.timestamp_format = timestamp_format
    
    def _render_generate_button(self, validation_results: Dict, original_data: pd.DataFrame):
        """Render the generate button and handle generation"""
        st.markdown("#### 🚀 Generate Report")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            if st.button("🔍 Generate Failed Records Report", type="primary", use_container_width=True):
                with st.spinner("Generating failed records report..."):
                    try:
                        failed_records_data = self._generate_failed_records_report(
                            validation_results, original_data
                        )
                        
                        if failed_records_data:
                            st.session_state.failed_records_data = failed_records_data
                            st.success("✅ Failed records report generated successfully!")
                            st.rerun()
                        else:
                            st.error("❌ Failed to generate report. No failed records found.")
                    except Exception as e:
                        st.error(f"❌ Error generating report: {str(e)}")
                        st.exception(e)
        
        with col2:
            if 'failed_records_data' in st.session_state and st.session_state.failed_records_data:
                # Download button
                self._render_download_button()
    
    def _render_download_button(self):
        """Render download button for generated report"""
        failed_records_data = st.session_state.failed_records_data
        report_format = st.session_state.get('failed_records_format', 'CSV')
        
        if not failed_records_data:
            return
        
        # Generate the file content
        file_content, file_name, mime_type = self._prepare_download_file(failed_records_data, report_format)
        
        if file_content:
            st.download_button(
                label=f"📥 Download {report_format} Report",
                data=file_content,
                file_name=file_name,
                mime=mime_type,
                use_container_width=True,
                key="download_failed_records_btn"
            )
    
    def _render_failed_records_preview(self):
        """Render preview of generated failed records data"""
        if 'failed_records_data' not in st.session_state:
            return
        
        failed_records_data = st.session_state.failed_records_data
        
        st.markdown("#### 👀 Report Preview")
        
        # Show summary
        total_records = len(failed_records_data['failed_records'])
        st.info(f"Generated report contains {total_records:,} failed records across {len(failed_records_data['expectation_summary'])} expectations.")
        
        # Show sample of failed records
        if failed_records_data['failed_records']:
            st.markdown("**Sample Failed Records:**")
            
            # Create a preview dataframe
            preview_data = []
            for record in failed_records_data['failed_records'][:10]:  # Show first 10
                preview_data.append({
                    'Row Index': record.get('row_index', 'N/A'),
                    'Failed Expectations': ', '.join(record.get('failed_expectations', [])),
                    'Primary Column': record.get('primary_column', 'N/A'),
                    'Failure Details': record.get('failure_details', 'N/A')[:100] + '...' if len(str(record.get('failure_details', ''))) > 100 else record.get('failure_details', 'N/A')
                })
            
            if preview_data:
                preview_df = pd.DataFrame(preview_data)
                st.dataframe(
                    preview_df,
                    use_container_width=True,
                    height=300
                )
        
        # Show expectation summary
        if failed_records_data['expectation_summary']:
            st.markdown("**Expectation Summary:**")
            summary_df = pd.DataFrame(failed_records_data['expectation_summary'])
            st.dataframe(
                summary_df,
                use_container_width=True,
                column_config={
                    "Expectation Type": st.column_config.TextColumn("Expectation Type", width="medium"),
                    "Column": st.column_config.TextColumn("Column", width="small"),
                    "Failed Records": st.column_config.NumberColumn("Failed Records", width="small"),
                    "Failure Rate": st.column_config.TextColumn("Failure Rate", width="small")
                }
            )
    
    def _calculate_failed_records_stats(self, validation_results: Dict, original_data: pd.DataFrame) -> Dict:
        """Calculate statistics about failed records"""
        try:
            results = validation_results.get('results', [])
            total_failed_records = 0
            failed_expectations = 0
            affected_columns = set()
            breakdown_by_type = []
            
            for result in results:
                if not result.get('success', False):
                    failed_expectations += 1
                    
                    exp_config = result.get('expectation_config', {})
                    exp_type = exp_config.get('type', exp_config.get('expectation_type', 'Unknown'))
                    column = exp_config.get('kwargs', {}).get('column', 'N/A')
                    
                    if column != 'N/A':
                        affected_columns.add(column)
                    
                    # Calculate failed records for this expectation
                    if 'result' in result and result['result']:
                        result_data = result['result']
                        unexpected_count = result_data.get('unexpected_count', 0)
                        missing_count = result_data.get('missing_count', 0)
                        failed_count = unexpected_count + missing_count
                        
                        if failed_count > 0:
                            total_failed_records += failed_count
                            
                            # Calculate failure rate
                            element_count = result_data.get('element_count', 0)
                            failure_rate = (failed_count / element_count * 100) if element_count > 0 else 0
                            
                            breakdown_by_type.append({
                                'Column': column,
                                'Failed Records': failed_count,
                                'Failure Rate': f"{failure_rate:.1f}%"
                            })
            
            return {
                'total_failed_records': total_failed_records,
                'failed_expectations': failed_expectations,
                'affected_columns': len(affected_columns),
                'breakdown_by_type': breakdown_by_type
            }
            
        except Exception as e:
            st.error(f"Error calculating failed records stats: {str(e)}")
            return {
                'total_failed_records': 0,
                'failed_expectations': 0,
                'affected_columns': 0,
                'breakdown_by_type': []
            }
    
    def _generate_failed_records_report(self, validation_results: Dict, original_data: pd.DataFrame) -> Dict:
        """Generate the complete failed records report"""
        try:
            results = validation_results.get('results', [])
            failed_records = []
            expectation_summary = []
            
            # Process each failed expectation
            for result in results:
                if not result.get('success', False):
                    exp_config = result.get('expectation_config', {})
                    exp_type = exp_config.get('type', exp_config.get('expectation_type', 'Unknown'))
                    column = exp_config.get('kwargs', {}).get('column', 'N/A')
                    
                    # Get failure details
                    failure_details = self._extract_failure_details(result, original_data)
                    
                    # Add to failed records
                    for record in failure_details:
                        failed_records.append(record)
                    
                    # Add to expectation summary
                    if failure_details:
                        expectation_summary.append({
                            'Column': column,
                            'Failed Records': len(failure_details),
                            'Failure Rate': f"{(len(failure_details) / len(original_data) * 100):.1f}%"
                        })
            
            if not failed_records:
                return None
            
            # Create the complete report
            report_data = {
                'metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'total_records_analyzed': len(original_data),
                    'total_failed_records': len(failed_records),
                    'validation_suite': st.session_state.get('current_suite_name', 'Unknown'),
                    'generation_options': {
                        'include_metadata': st.session_state.get('include_metadata', True),
                        'include_original': st.session_state.get('include_original', True),
                        'group_by_expectation': st.session_state.get('group_by_expectation', False),
                        'timestamp_format': st.session_state.get('timestamp_format', 'ISO')
                    }
                },
                'failed_records': failed_records,
                'expectation_summary': expectation_summary,
                'success_summary': self._generate_success_summary(validation_results) if st.session_state.get('include_success_summary', True) else None
            }
            
            return report_data
            
        except Exception as e:
            st.error(f"Error generating failed records report: {str(e)}")
            return None
    
    def _extract_failure_details(self, result: Dict, original_data: pd.DataFrame) -> List[Dict]:
        """Extract detailed failure information for a specific expectation result"""
        try:
            failure_details = []
            exp_config = result.get('expectation_config', {})
            exp_type = exp_config.get('type', exp_config.get('expectation_type', 'Unknown'))
            column = exp_config.get('kwargs', {}).get('column', 'N/A')
            
            if 'result' not in result or not result['result']:
                return failure_details
            
            result_data = result['result']
            
            # Handle different types of failures
            if 'unexpected_list' in result_data and result_data['unexpected_list']:
                # Records that failed the expectation
                for unexpected_value in result_data['unexpected_list']:
                    # Find rows with this value
                    if column != 'N/A' and column in original_data.columns:
                        failed_rows = original_data[original_data[column] == unexpected_value]
                        
                        for idx, row in failed_rows.iterrows():
                            failure_details.append({
                                'row_index': idx,
                                'failed_expectations': [exp_type],
                                'primary_column': column,
                                'failed_value': unexpected_value,
                                'expectation_type': exp_type,
                                'failure_reason': f"Value '{unexpected_value}' failed {exp_type}",
                                'failure_details': self._format_failure_details(result_data, exp_config),
                                'original_record': row.to_dict() if st.session_state.get('include_original', True) else None,
                                'metadata': self._extract_metadata(result, exp_config) if st.session_state.get('include_metadata', True) else None
                            })
            
            elif 'missing_list' in result_data and result_data['missing_list']:
                # Records that are missing (for completeness expectations)
                for missing_value in result_data['missing_list']:
                    failure_details.append({
                        'row_index': 'N/A',
                        'failed_expectations': [exp_type],
                        'primary_column': column,
                        'failed_value': missing_value,
                        'expectation_type': exp_type,
                        'failure_reason': f"Missing value '{missing_value}' for {exp_type}",
                        'failure_details': self._format_failure_details(result_data, exp_config),
                        'original_record': None,
                        'metadata': self._extract_metadata(result, exp_config) if st.session_state.get('include_metadata', True) else None
                    })
            
            # If no specific unexpected/missing lists, create general failure record
            if not failure_details and not result.get('success', False):
                failure_details.append({
                    'row_index': 'N/A',
                    'failed_expectations': [exp_type],
                    'primary_column': column,
                    'failed_value': 'N/A',
                    'expectation_type': exp_type,
                    'failure_reason': f"Expectation {exp_type} failed validation",
                    'failure_details': self._format_failure_details(result_data, exp_config),
                    'original_record': None,
                    'metadata': self._extract_metadata(result, exp_config) if st.session_state.get('include_metadata', True) else None
                })
            
            return failure_details
            
        except Exception as e:
            st.error(f"Error extracting failure details: {str(e)}")
            return []
    
    def _format_failure_details(self, result_data: Dict, exp_config: Dict) -> str:
        """Format failure details in a readable way"""
        try:
            details = []
            
            if 'unexpected_count' in result_data:
                details.append(f"Unexpected records: {result_data['unexpected_count']}")
            
            if 'missing_count' in result_data:
                details.append(f"Missing records: {result_data['missing_count']}")
            
            if 'unexpected_percent' in result_data:
                details.append(f"Unexpected percentage: {result_data['unexpected_percent']:.1f}%")
            
            if 'missing_percent' in result_data:
                details.append(f"Missing percentage: {result_data['missing_percent']:.1f}%")
            
            if 'partial_unexpected_list' in result_data and result_data['partial_unexpected_list']:
                sample_values = result_data['partial_unexpected_list'][:5]  # Show first 5
                details.append(f"Sample unexpected values: {sample_values}")
            
            return ' | '.join(details) if details else "No detailed failure information available"
            
        except Exception as e:
            return f"Error formatting details: {str(e)}"
    
    def _extract_metadata(self, result: Dict, exp_config: Dict) -> Dict:
        """Extract metadata about the expectation and validation"""
        try:
            metadata = {
                'expectation_type': exp_config.get('type', exp_config.get('expectation_type', 'Unknown')),
                'expectation_kwargs': exp_config.get('kwargs', {}),
                'validation_timestamp': result.get('meta', {}).get('run_time', datetime.now().isoformat()),
                'result_keys': list(result.keys()) if result else []
            }
            
            # Add specific result metadata
            if 'result' in result and result['result']:
                result_data = result['result']
                metadata.update({
                    'element_count': result_data.get('element_count', 0),
                    'unexpected_count': result_data.get('unexpected_count', 0),
                    'missing_count': result_data.get('missing_count', 0)
                })
            
            return metadata
            
        except Exception as e:
            return {'error': f"Error extracting metadata: {str(e)}"}
    
    def _generate_success_summary(self, validation_results: Dict) -> Dict:
        """Generate summary of successful expectations"""
        try:
            results = validation_results.get('results', [])
            successful_expectations = []
            
            for result in results:
                if result.get('success', False):
                    exp_config = result.get('expectation_config', {})
                    exp_type = exp_config.get('type', exp_config.get('expectation_type', 'Unknown'))
                    column = exp_config.get('kwargs', {}).get('column', 'N/A')
                    
                    successful_expectations.append({
                        'expectation_type': exp_type.replace('expect_', '').replace('_', ' ').title(),
                        'column': column,
                        'status': 'Passed',
                        'details': 'Expectation passed successfully'
                    })
            
            return {
                'total_passed': len(successful_expectations),
                'expectations': successful_expectations
            }
            
        except Exception as e:
            return {'error': f"Error generating success summary: {str(e)}"}
    
    def _prepare_download_file(self, failed_records_data: Dict, report_format: str) -> Tuple[Any, str, str]:
        """Prepare the file content for download"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            suite_name = st.session_state.get('current_suite_name', 'validation_suite')
            
            if report_format == 'CSV':
                # Convert to CSV format
                csv_data = self._convert_to_csv(failed_records_data)
                file_name = f"failed_records_{suite_name}_{timestamp}.csv"
                mime_type = "text/csv"
                return csv_data, file_name, mime_type
                
            elif report_format == 'Excel':
                # Convert to Excel format
                excel_data = self._convert_to_excel(failed_records_data)
                file_name = f"failed_records_{suite_name}_{timestamp}.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                return excel_data, file_name, mime_type
                
            elif report_format == 'JSON':
                # Convert to JSON format
                json_data = self._convert_to_json(failed_records_data)
                file_name = f"failed_records_{suite_name}_{timestamp}.json"
                mime_type = "application/json"
                return json_data, file_name, mime_type
            
            return None, "", ""
            
        except Exception as e:
            st.error(f"Error preparing download file: {str(e)}")
            return None, "", ""
    
    def _convert_to_csv(self, failed_records_data: Dict) -> str:
        """Convert failed records data to CSV format"""
        try:
            # Create a flattened CSV structure
            csv_rows = []
            
            # Add header
            headers = ['Row Index', 'Failed Expectations', 'Primary Column', 'Failed Value', 
                      'Expectation Type', 'Failure Reason', 'Failure Details']
            
            if st.session_state.get('include_original', True):
                # Add original data columns
                if failed_records_data['failed_records'] and failed_records_data['failed_records'][0].get('original_record'):
                    original_columns = list(failed_records_data['failed_records'][0]['original_record'].keys())
                    headers.extend([f"Original_{col}" for col in original_columns])
            
            csv_rows.append(','.join(headers))
            
            # Add data rows
            for record in failed_records_data['failed_records']:
                row_data = [
                    str(record.get('row_index', 'N/A')),
                    '; '.join(record.get('failed_expectations', [])),
                    str(record.get('primary_column', 'N/A')),
                    str(record.get('failed_value', 'N/A')),
                    str(record.get('expectation_type', 'N/A')),
                    str(record.get('failure_reason', 'N/A')),
                    str(record.get('failure_details', 'N/A'))
                ]
                
                if st.session_state.get('include_original', True) and record.get('original_record'):
                    for value in record['original_record'].values():
                        row_data.append(str(value))
                
                # Escape quotes in CSV values
                escaped_values = []
                for val in row_data:
                    escaped_val = str(val).replace('"', '""')
                    escaped_values.append(f'"{escaped_val}"')
                csv_rows.append(','.join(escaped_values))
            
            return '\n'.join(csv_rows)
            
        except Exception as e:
            st.error(f"Error converting to CSV: {str(e)}")
            return ""
    
    def _convert_to_excel(self, failed_records_data: Dict) -> bytes:
        """Convert failed records data to Excel format"""
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.title = "Summary"
            
            # Add summary data
            summary_ws['A1'] = "Failed Records Report Summary"
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            summary_ws['A3'] = "Generated At:"
            summary_ws['B3'] = failed_records_data['metadata']['generated_at']
            
            summary_ws['A4'] = "Total Records Analyzed:"
            summary_ws['B4'] = failed_records_data['metadata']['total_records_analyzed']
            
            summary_ws['A5'] = "Total Failed Records:"
            summary_ws['B5'] = failed_records_data['metadata']['total_failed_records']
            
            # Add expectation summary
            if failed_records_data['expectation_summary']:
                summary_ws['A7'] = "Expectation Summary"
                summary_ws['A7'].font = Font(bold=True, size=12)
                
                # Headers
                headers = ['Column', 'Failed Records', 'Failure Rate']
                for col, header in enumerate(headers, 1):
                    cell = summary_ws.cell(row=8, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Data
                for row, summary in enumerate(failed_records_data['expectation_summary'], 9):
                    summary_ws.cell(row=row, column=1, value=summary['Column'])
                    summary_ws.cell(row=row, column=2, value=summary['Failed Records'])
                    summary_ws.cell(row=row, column=3, value=summary['Failure Rate'])
            
            # Create failed records sheet
            records_ws = wb.create_sheet("Failed Records")
            
            # Add headers
            headers = ['Row Index', 'Failed Expectations', 'Primary Column', 'Failed Value', 
                      'Expectation Type', 'Failure Reason', 'Failure Details']
            
            if st.session_state.get('include_original', True):
                # Add original data columns
                if failed_records_data['failed_records'] and failed_records_data['failed_records'][0].get('original_record'):
                    original_columns = list(failed_records_data['failed_records'][0]['original_record'].keys())
                    headers.extend([f"Original_{col}" for col in original_columns])
            
            for col, header in enumerate(headers, 1):
                cell = records_ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row, record in enumerate(failed_records_data['failed_records'], 2):
                col = 1
                records_ws.cell(row=row, column=col, value=str(record.get('row_index', 'N/A'))); col += 1
                records_ws.cell(row=row, column=col, value='; '.join(record.get('failed_expectations', []))); col += 1
                records_ws.cell(row=row, column=col, value=str(record.get('primary_column', 'N/A'))); col += 1
                records_ws.cell(row=row, column=col, value=str(record.get('failed_value', 'N/A'))); col += 1
                records_ws.cell(row=row, column=col, value=str(record.get('expectation_type', 'N/A'))); col += 1
                records_ws.cell(row=row, column=col, value=str(record.get('failure_reason', 'N/A'))); col += 1
                records_ws.cell(row=row, column=col, value=str(record.get('failure_details', 'N/A'))); col += 1
                
                if st.session_state.get('include_original', True) and record.get('original_record'):
                    for value in record['original_record'].values():
                        records_ws.cell(row=row, column=col, value=str(value))
                        col += 1
            
            # Auto-adjust column widths
            for ws in [summary_ws, records_ws]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error converting to Excel: {str(e)}")
            return b""
    
    def _convert_to_json(self, failed_records_data: Dict) -> str:
        """Convert failed records data to JSON format"""
        try:
            # Format timestamps if needed
            formatted_data = self._format_timestamps(failed_records_data)
            
            return json.dumps(formatted_data, indent=2, default=str)
            
        except Exception as e:
            st.error(f"Error converting to JSON: {str(e)}")
            return "{}"
    
    def _format_timestamps(self, data: Dict) -> Dict:
        """Format timestamps according to user preference"""
        try:
            timestamp_format = st.session_state.get('timestamp_format', 'ISO')
            
            if timestamp_format == 'Readable':
                # Convert ISO timestamps to readable format
                if 'metadata' in data and 'generated_at' in data['metadata']:
                    try:
                        dt = datetime.fromisoformat(data['metadata']['generated_at'])
                        data['metadata']['generated_at'] = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                
                # Format timestamps in metadata
                for record in data.get('failed_records', []):
                    if 'metadata' in record and 'validation_timestamp' in record['metadata']:
                        try:
                            dt = datetime.fromisoformat(record['metadata']['validation_timestamp'])
                            record['metadata']['validation_timestamp'] = dt.strftime('%Y-%m-%d %H:%M:%S')
                        except:
                            pass
            
            elif timestamp_format == 'Unix':
                # Convert to Unix timestamps
                if 'metadata' in data and 'generated_at' in data['metadata']:
                    try:
                        dt = datetime.fromisoformat(data['metadata']['generated_at'])
                        data['metadata']['generated_at'] = int(dt.timestamp())
                    except:
                        pass
                
                # Convert timestamps in metadata
                for record in data.get('failed_records', []):
                    if 'metadata' in record and 'validation_timestamp' in record['metadata']:
                        try:
                            dt = datetime.fromisoformat(record['metadata']['validation_timestamp'])
                            record['metadata']['validation_timestamp'] = int(dt.timestamp())
                        except:
                            pass
            
            return data
            
        except Exception as e:
            return data
